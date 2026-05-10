from pathlib import Path
import sys

try:
    from openpyxl import load_workbook
except Exception as exc:
    print(f"❌ openpyxl unavailable: {exc}")
    sys.exit(1)

root = Path(__file__).resolve().parents[1]
xlsx = root / "templates" / "compliantuk-portfolio-template.xlsx"
expected = [
    "landlord_first", "landlord_last", "landlord_email", "property_address",
    "tenant1_first", "tenant1_last", "tenant1_email",
    "tenant2_first", "tenant2_last", "tenant2_email",
    "tenant3_first", "tenant3_last", "tenant3_email",
    "tenant4_first", "tenant4_last", "tenant4_email",
    "tenant5_first", "tenant5_last", "tenant5_email",
    "tenant6_first", "tenant6_last", "tenant6_email",
]

passed = 0
failed = 0

def check(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  ✅ {name}")
        passed += 1
    else:
        print(f"  ❌ {name}: {detail}")
        failed += 1

print("\n📊 XLSX TEMPLATE INTEGRITY TESTS")
check("XLSX template exists", xlsx.exists(), str(xlsx))
if xlsx.exists():
    wb = load_workbook(xlsx)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    check("XLSX has exact expected headers through tenant6", headers[: len(expected)] == expected, str(headers[: len(expected)]))
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    processable = [row for row in rows if row and row[3] and row[4] and row[6]]
    check("XLSX has processable sample rows", len(processable) >= 2, f"processable={len(processable)}")
    first = rows[0] if rows else []
    tenant_emails = [first[i] for i in range(6, min(len(first), 22), 3) if first[i]] if first else []
    check("XLSX sample demonstrates more than four tenants", len(tenant_emails) > 4, f"tenant_emails={tenant_emails}")
    check("XLSX includes Instructions sheet", "Instructions" in wb.sheetnames, str(wb.sheetnames))

print("=" * 60)
print(f"XLSX TEMPLATE TEST SUMMARY: {passed} passed, {failed} failed")
print("=" * 60)
sys.exit(1 if failed else 0)
